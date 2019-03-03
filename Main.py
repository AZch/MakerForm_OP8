import datetime
import sys

from PyQt5 import QtWidgets
from PyQt5 import QtCore
from openpyxl import *
from openpyxl.styles import Font, Alignment

from Constants import *

import lr3


class MainWnd(QtWidgets.QMainWindow, lr3.Ui_MainWindow):
    arrElemName = ["None", nameInTbl.spoon, nameInTbl.fork, nameInTbl.glass, nameInTbl.cup, nameInTbl.plate]
    arrElemOrg = ["None", nameOrg.CompanyPleasantDishes, nameOrg.CompanyRefinedSilkOfTheEast, nameOrg.Enterprise2IsVictorious, nameOrg.SchoolСanteen127]
    arrElemStruct = ["None", nameStruct.CafeСomfort, nameStruct.Canteen, nameStruct.ShoppingRoom, nameStruct.Stock, nameStruct.TeachingStaff]
    maxSpnTbl = 999999999

    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.initUI()

    def initUI(self):
        self.tabMain.mouseReleaseEvent = self.__updateMainTblRow()
        self.tabFight.mouseReleaseEvent = self.__updateTblFight
        self.tabLost.mouseReleaseEvent = self.__updateTblLost

        self.cmbOrg.addItems(self.arrElemOrg)
        self.cmbOrg.activated[str].connect(self.changeOrg)

        self.cmbStruct.addItems(self.arrElemStruct)
        self.cmbStruct.activated[str].connect(self.changeStruct)

        self.exit.clicked.connect(self.__closeApp)
        self.exportExcel.clicked.connect(self.__exportData)

    def __closeApp(self):
        qMessBox = QtWidgets.QMessageBox
        if qMessBox.question(self, 'Закрытие', 'Вы можете потерять данные, вы уверены?',
                             qMessBox.Yes | qMessBox.No) == qMessBox.Yes:
            QtCore.QCoreApplication.instance().quit()

    def __cellCenter(self, text, cell, font, ws):
        self.__cellSimple(text, cell, font, ws)
        ws[cell].alignment = Alignment(horizontal="center", vertical='center')

    def __cellLeft(self, text, cell, font, ws):
        self.__cellSimple(text, cell, font, ws)
        ws[cell].alignment = Alignment(horizontal="left", vertical='top')

    def __cellSimple(self, text, cell, font, ws):
        ws[cell] = text
        ws[cell].font = Font(size=font)

    def MakeRow(self, row, dataRow, ws):
        ws.merge_cells('A' + str(row) + ':D' + str(row))
        self.__cellCenter(dataRow[0], 'A' + str(row), 10, ws)

        ws.merge_cells('E' + str(row) + ':N' + str(row))
        self.__cellCenter(dataRow[1], 'E' + str(row), 10, ws)

        ws.merge_cells('O' + str(row) + ':S' + str(row))
        self.__cellCenter(dataRow[2], 'O' + str(row), 10, ws)

        ws.merge_cells('T' + str(row) + ':Y' + str(row))
        self.__cellCenter(dataRow[3], 'T' + str(row), 10, ws)

        ws.merge_cells('Z' + str(row) + ':AC' + str(row))
        self.__cellCenter(dataRow[4], 'Z' + str(row), 10, ws)

        ws.merge_cells('AD' + str(row) + ':AH' + str(row))
        self.__cellCenter(dataRow[5], 'AD' + str(row), 10, ws)

        ws.merge_cells('AI' + str(row) + ':AL' + str(row))
        self.__cellCenter(dataRow[6], 'AI' + str(row), 10, ws)

        ws.merge_cells('AM' + str(row) + ':AQ' + str(row))
        self.__cellCenter(dataRow[7], 'AM' + str(row), 10, ws)

        ws.merge_cells('AR' + str(row) + ':AU' + str(row))
        self.__cellCenter(dataRow[8], 'AR' + str(row), 10, ws)

        ws.merge_cells('AV' + str(row) + ':AZ' + str(row))
        self.__cellCenter(dataRow[9], 'AV' + str(row), 10, ws)

        ws.merge_cells('BA' + str(row) + ':BQ' + str(row))
        self.__cellCenter(dataRow[10], 'BA' + str(row), 10, ws)

        ws.merge_cells('BR' + str(row) + ':BX' + str(row))
        self.__cellCenter(dataRow[11], 'BR' + str(row), 10, ws)

    def __allFightCount(self, rowStart, rowEnd):
        i = rowStart
        count = 0
        while i < rowEnd:
            count += self.tblFight.cellWidget(i, 1).value()
            i += 1
        return count

    def __allFightPrice(self, rowStart, rowEnd):
        i = rowStart
        count = 0
        while i < rowEnd:
            count += int(self.tblFight.item(i, 2).text())
            i += 1
        return count

    def __allLostCount(self, rowStart, rowEnd):
        i = rowStart
        count = 0
        while i < rowEnd:
            count += self.tblLost.cellWidget(i, 1).value()
            i += 1
        return count

    def __allLostPrice(self, rowStart, rowEnd):
        i = rowStart
        count = 0
        while i < rowEnd:
            count += int(self.tblLost.item(i, 2).text())
            i += 1
        return count

    def __allLostFightCount(self, rowStart, rowEnd):
        return self.__allFightCount(rowStart, rowEnd) + self.__allLostCount(rowStart, rowEnd)

    def __allLostFightPrice(self, rowStart, rowEnd):
        return self.__allFightPrice(rowStart, rowEnd) + self.__allLostPrice(rowStart, rowEnd)

    def __fillTblEmptyForm(self, strStar, strEnd, ws):
        while strStar < strEnd:
            self.MakeRow(strStar, [
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ], ws)
            strStar += 1

    def __exportData(self):
        #wb = load_workbook(filename = 'forma-op-8.xlsx')
        wb = Workbook()
        ws = wb.active
        self.__cellSimple("Унифицированная форма № ОП-8", 'BG1', 9, ws)
        self.__cellSimple("Утверждена постановлением Госкомстата", 'BG2', 9, ws)
        self.__cellSimple("России от 25.12.98 № 132", 'BG3', 9, ws)
        self.__cellCenter("КОД", 'BQ4', 10, ws)
        i = 4
        while i < 11:
            strMerge = 'BQ' + str(i) + ':BX' + str(i)
            ws.merge_cells(strMerge)
            i += 1
        ws.merge_cells('A6:BJ6')
        ws.merge_cells('A8:BP8')
        ws.merge_cells('BM11:BW11')
        ws.merge_cells('BM12:BW12')
        ws.merge_cells('BM13:BW13')
        ws.merge_cells('BM14:BW14')
        ws.merge_cells('BL15:BO15')
        ws.merge_cells('BQ15:BX15')
        ws.merge_cells('BL16:BO16')
        ws.merge_cells('BQ16:BX16')
        ws.merge_cells('BN17:BT17')
        ws.merge_cells('BV17:BW17')
        ws.merge_cells('U18:AK18')
        ws.merge_cells('AM18:BY18')
        ws.merge_cells('U19:AK19')
        ws.merge_cells('AM19:BY19')

        ws.merge_cells('AK12:AR13')
        ws.merge_cells('AK14:AR15')
        ws.merge_cells('AS12:AZ13')
        ws.merge_cells('AS14:AZ15')
        ws.merge_cells('BA12:BJ12')
        ws.merge_cells('BA13:BE13')
        ws.merge_cells('BA14:BE15')
        ws.merge_cells('BF13:BJ13')
        ws.merge_cells('BF14:BJ15')
        ws.merge_cells('AG14:AJ15')

        ws.merge_cells('A21:D26')
        ws.merge_cells('E21:S22')
        ws.merge_cells('E23:N26')
        ws.merge_cells('O23:S26')
        ws.merge_cells('T21:Y26')
        ws.merge_cells('Z21:AZ22')
        ws.merge_cells('Z23:AH23')
        ws.merge_cells('AI23:AQ23')
        ws.merge_cells('AR23:AZ23')
        ws.merge_cells('Z24:AC26')
        ws.merge_cells('AD24:AH26')
        ws.merge_cells('AI24:AL26')
        ws.merge_cells('AM24:AQ26')
        ws.merge_cells('AR24:AU26')
        ws.merge_cells('AV24:AZ26')
        ws.merge_cells('BA21:BQ26')
        ws.merge_cells('BR21:BX26')
        self.__cellCenter("Но-\nмер\nпо по-\nрядку", "A21", 10, ws)
        self.__cellCenter("Посуда, приборы", "E21", 10, ws)
        self.__cellCenter("наименование", "E23", 10, ws)
        self.__cellCenter("код", "O23", 10, ws)
        self.__cellCenter("Цена,\nруб. коп.", "T21", 10, ws)
        self.__cellCenter("Бой, лом, утрачено, пропало", "Z21", 10, ws)
        self.__cellCenter("бой, лом ", "Z23", 10, ws)
        self.__cellCenter("утрачено, пропало", "AI23", 10, ws)
        self.__cellCenter("всего", "AR23", 10, ws)
        self.__cellCenter("коли-\nчество,\nшт.", "Z24", 10, ws)
        self.__cellCenter("сумма,\nруб. коп.", "AD24", 10, ws)
        self.__cellCenter("коли-\nчество,\nшт.", "AI24", 10, ws)
        self.__cellCenter("сумма,\nруб. коп.", "AM24", 10, ws)
        self.__cellCenter("коли-\nчество,\nшт.", "AR24", 10, ws)
        self.__cellCenter("сумма,\nруб. коп.", "AV24", 10, ws)
        self.__cellCenter("Обстоятельства\nбоя, лома, утраты, пропажи.\nВиновные лица\n(должность, фамилия, и., о.)", "BA21", 10, ws)
        self.__cellCenter("Примечание", "BR21", 10, ws)

        self.__cellCenter("0330508", 'BQ5', 10, ws)
        self.__cellSimple("Форма по ОКУД", 'BH5', 10, ws)
        self.__cellSimple("по ОКПО", 'BL6', 10, ws)
        self.__cellSimple("Вид деятельности по ОКДП", 'BC9', 10, ws)
        self.__cellSimple("Вид операции", 'BI10', 10, ws)
        self.__cellSimple("(организация)", 'AC7', 7, ws)
        self.__cellSimple("(структурное подразделение)", 'AD9', 7, ws)

        self.__cellCenter("АКТ", 'AG14', 15, ws)
        self.__cellCenter("Номер\nДокумента", 'AK12', 9, ws)
        self.__cellSimple("О БОЕ, ЛОМЕ И УТРАТЕ ПОСУДЫ И ПРИБОРОВ", 'U16', 11, ws)
        self.__cellSimple("Материально ответственное лицо", 'C18', 11, ws)
        self.__cellSimple("Комиссия установила:", 'C20', 11, ws)
        self.__cellCenter("(должность)", 'U19', 8, ws)
        self.__cellCenter("(фамилия, имя, отчество)", 'AM19', 8, ws)
        self.__cellCenter("УТВЕРЖДАЮ", 'BM11', 9, ws)
        self.__cellCenter("Руководитель", 'BM12', 10, ws)
        self.__cellCenter("(должность)", 'BM14', 6.5, ws)
        self.__cellCenter("(подпись)", 'BL16', 6.5, ws)
        self.__cellCenter("(расшифровка подписи)", 'BQ16', 6.5, ws)
        self.__cellCenter("<<", 'BK17', 5, ws)
        self.__cellCenter(">>", 'BM17', 5, ws)
        self.__cellCenter("г.", 'BX17', 10, ws)
        self.__cellCenter("с", 'BA13', 9, ws)
        self.__cellCenter("по", 'BF13', 9, ws)
        self.__cellCenter("Дата\nсоставления", 'AS12', 9, ws)
        self.__cellCenter("Отчетный период", 'BA12', 9, ws)
        i = ord('A') - 1
        j = ord('A')
        while (not (i == ord('B') and j == ord('X') + 1)):
            strCell = ''
            if i == ord('A') - 1:
                strCell = chr(j)
            if j == ord('Z') + 1:
                i += 1
                j = ord('A')
                strCell = chr(i) + chr(j)
            elif i != ord('A') - 1:
                strCell = chr(i) + chr(j)

            ws.column_dimensions[strCell].width = 1.77
            j += 1
        # for k in range(20):
        #     ws.row_dimensions[str(k + 1)].height = 3

        self.__cellCenter(self.cmbOrg.currentText(), "A6", 11, ws)
        self.__cellCenter(self.cmbStruct.currentText(), "A8", 11, ws)
        self.__cellCenter(str(self.numOrg.value()), "BQ6", 10, ws)
        self.__cellCenter(str(self.numStruct.value()), "BQ8", 10, ws)
        self.__cellCenter(str(self.typeOperation.value()), "BQ10", 10, ws)
        self.__cellCenter(self.headPosition.text(), "BM13", 8, ws)
        self.__cellCenter(self.headFullName.text(), "BQ15", 6.5, ws)
        self.__cellCenter(self.numAct.text(), "AK14", 9, ws)
        self.__cellCenter(self.dateEditAct.text(), "AS14", 9, ws)
        self.__cellCenter(self.reportForm.text(), "BA14", 9, ws)
        self.__cellCenter(self.reportOn.text(), "BF14", 9, ws)

        self.__cellCenter(self.dateCreate.text().split(".")[0], "BL17", 7, ws)
        self.__cellCenter(self.dateCreate.text().split(".")[1], "BN17", 7, ws)
        self.__cellCenter(self.dateCreate.text().split(".")[2], "BV17", 7, ws)

        self.__cellCenter(self.position.text(), "U18", 11, ws)
        self.__cellCenter(self.FIO.text(), "AM18", 11, ws)

        self.MakeRow(27, [
            str(1),
            str(2),
            str(3),
            str(4),
            str(5),
            str(6),
            str(7),
            str(8),
            str(9),
            str(10),
            str(11),
            str(12),
        ], ws)

        rowStart = 28

        baseDiv = 10
        rowDiv = baseDiv
        if self.tblMain.rowCount() > rowDiv * 2:
            rowDiv = int(self.tblMain.rowCount() / 2)
        elif rowDiv >= self.tblMain.rowCount():
            rowDiv = self.tblMain.rowCount()



        for row in range(rowDiv):
            self.MakeRow(rowStart + row, [
                str(row + 1),
                self.tblMain.cellWidget(row, 0).currentText(),
                str(self.tblMain.cellWidget(row, 1).value()),
                str(self.tblMain.cellWidget(row, 2).value()),
                str(self.tblFight.cellWidget(row, 1).value()),
                str(self.tblFight.item(row, 2).text()),
                str(self.tblLost.cellWidget(row, 1).value()),
                str(self.tblLost.item(row, 2).text()),
                str(self.tblLost.item(row, 4).text()),
                str(self.tblLost.item(row, 5).text()),
                str(self.tblFight.cellWidget(row, 3).toPlainText()) + ", " + str(self.tblLost.cellWidget(row, 3).toPlainText()),
                str(self.tblMain.cellWidget(row, 3).toPlainText()),
            ], ws)

        if rowDiv < baseDiv:
            self.__fillTblEmptyForm(rowStart + rowDiv, rowStart + baseDiv, ws)

        self.MakeRow(rowStart + baseDiv, [
            "",
            "",
            "",
            "Итого",
            str(self.__allFightCount(0, rowDiv)),
            str(self.__allFightPrice(0, rowDiv)),
            str(self.__allLostCount(0, rowDiv)),
            str(self.__allLostPrice(0, rowDiv)),
            str(self.__allLostFightCount(0, rowDiv)),
            str(self.__allLostFightPrice(0, rowDiv)),
            "",
            "",
        ], ws)

        row = rowStart + baseDiv + 1

        ws.merge_cells('BE' + str(row) + ':BX' + str(row))
        self.__cellCenter('Оборотная сторона формы № ОП-8', 'BE' + str(row), 11, ws)

        row += 2

        ws.merge_cells('A' + str(row) + ':D' + str(row + 5))
        ws.merge_cells('E' + str(row) + ':S' + str(row + 1))
        ws.merge_cells('E' + str(row + 2) + ':N' + str(row + 5))
        ws.merge_cells('O' + str(row + 2) + ':S' + str(row + 5))
        ws.merge_cells('T' + str(row) + ':Y' + str(row + 5))
        ws.merge_cells('Z' + str(row) +':AZ' + str(row + 1))
        ws.merge_cells('Z' + str(row + 2) + ':AH' + str(row + 2))
        ws.merge_cells('AI' + str(row + 2) + ':AQ' + str(row + 2))
        ws.merge_cells('AR' + str(row + 2) + ':AZ' + str(row + 2))
        ws.merge_cells('Z' + str(row + 3) + ':AC' + str(row + 5))
        ws.merge_cells('AD' + str(row + 3) + ':AH' + str(row + 5))
        ws.merge_cells('AI' + str(row + 3) + ':AL' + str(row + 5))
        ws.merge_cells('AM' + str(row + 3) + ':AQ' + str(row + 5))
        ws.merge_cells('AR' + str(row + 3) + ':AU' + str(row + 5))
        ws.merge_cells('AV' + str(row + 3) + ':AZ' + str(row + 5))
        ws.merge_cells('BA' + str(row) + ':BQ' + str(row + 5))
        ws.merge_cells('BR' + str(row) + ':BX' + str(row + 5))
        self.__cellCenter("Но-\nмер\nпо по-\nрядку", "A" + str(row), 10, ws)
        self.__cellCenter("Посуда, приборы", "E" + str(row), 10, ws)
        self.__cellCenter("наименование", "E" + str(row + 2), 10, ws)
        self.__cellCenter("код", "O" + str(row + 2), 10, ws)
        self.__cellCenter("Цена,\nруб. коп.", "T" + str(row), 10, ws)
        self.__cellCenter("Бой, лом, утрачено, пропало", "Z" + str(row), 10, ws)
        self.__cellCenter("бой, лом ", "Z" + str(row + 2), 10, ws)
        self.__cellCenter("утрачено, пропало", "AI" + str(row + 2), 10, ws)
        self.__cellCenter("всего", "AR" + str(row + 2), 10, ws)
        self.__cellCenter("коли-\nчество,\nшт.", "Z" + str(row + 3), 10, ws)
        self.__cellCenter("сумма,\nруб. коп.", "AD" + str(row + 3), 10, ws)
        self.__cellCenter("коли-\nчество,\nшт.", "AI" + str(row + 3), 10, ws)
        self.__cellCenter("сумма,\nруб. коп.", "AM" + str(row + 3), 10, ws)
        self.__cellCenter("коли-\nчество,\nшт.", "AR" + str(row + 3), 10, ws)
        self.__cellCenter("сумма,\nруб. коп.", "AV" + str(row + 3), 10, ws)
        self.__cellCenter("Обстоятельства\nбоя, лома, утраты, пропажи.\nВиновные лица\n(должность, фамилия, и., о.)",
                          "BA" + str(row), 10, ws)
        self.__cellCenter("Примечание", "BR" + str(row), 10, ws)

        row += 6
        self.MakeRow(row, [
            str(1),
            str(2),
            str(3),
            str(4),
            str(5),
            str(6),
            str(7),
            str(8),
            str(9),
            str(10),
            str(11),
            str(12),
        ], ws)
        row += 1
        i = self.tblMain.rowCount() - rowDiv
        startI = i
        endI = self.tblMain.rowCount()
        if i > 0:
            while i < endI:
                self.MakeRow(i + row - startI, [
                    str(i + 1),
                    self.tblMain.cellWidget(i, 0).currentText(),
                    str(self.tblMain.cellWidget(i, 1).value()),
                    str(self.tblMain.cellWidget(i, 2).value()),
                    str(self.tblFight.cellWidget(i, 1).value()),
                    str(self.tblFight.item(i, 2).text()),
                    str(self.tblLost.cellWidget(i, 1).value()),
                    str(self.tblLost.item(i, 2).text()),
                    str(self.tblLost.item(i, 4).text()),
                    str(self.tblLost.item(i, 5).text()),
                    str(self.tblFight.cellWidget(i, 3).toPlainText()) + ", " + str(
                        self.tblLost.cellWidget(i, 3).toPlainText()),
                    str(self.tblMain.cellWidget(i, 3).toPlainText()),
                ], ws)
                i += 1
            self.MakeRow(row + baseDiv, [
                "",
                "",
                "",
                "Итого",
                str(self.__allFightCount(startI, endI)),
                str(self.__allFightPrice(startI, endI)),
                str(self.__allLostCount(startI, endI)),
                str(self.__allLostPrice(startI, endI)),
                str(self.__allLostFightCount(startI, endI)),
                str(self.__allLostFightPrice(startI, endI)),
                "",
                "",
            ], ws)
        else:
            self.__fillTblEmptyForm(row, row  + baseDiv, ws)
            self.MakeRow(row + baseDiv, [
                "",
                "",
                "",
                "Итого",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ], ws)

        self.MakeRow(row + baseDiv + 1, [
            "",
            "",
            "",
            "Всего",
            str(self.__allFightCount(0, endI)),
            str(self.__allFightPrice(0, endI)),
            str(self.__allLostCount(0, endI)),
            str(self.__allLostPrice(0, endI)),
            str(self.__allLostFightCount(0, endI)),
            str(self.__allLostFightPrice(0, endI)),
            "",
            "",
        ], ws)

        row += baseDiv + 2
        self.__cellSimple("Перечисленные в графе «Бой, лом» столовая посуда и приборы в количестве", 'A' + str(row), 10, ws)
        ws.merge_cells('AI' + str(row) + ':BF' + str(row))
        self.__cellCenter(str(self.cntCrushSee.value()), 'AI' + str(row), 10, ws)
        self.__cellSimple("шт. уничтожены в нашем присутствии", 'BG' + str(row), 10, ws)
        row += 1
        ws.merge_cells('AI' + str(row) + ':BF' + str(row))
        self.__cellCenter("(печатно)", 'AI' + str(row), 10, ws)
        row += 2
        self.__cellCenter("Члены комиссии:", 'E' + str(row), 10, ws)
        row += 1
        for rowTbl in range(self.tblCommis.rowCount()):
            ws.merge_cells('N' + str(row) + ':AB' + str(row))
            self.__cellCenter(str(self.tblCommis.cellWidget(rowTbl, 0).toPlainText()), 'N' + str(row), 10, ws)
            ws.merge_cells('AD' + str(row) + ':AO' + str(row))
            ws.merge_cells('AQ' + str(row) + ':BO' + str(row))
            self.__cellCenter(str(self.tblCommis.cellWidget(rowTbl, 1).toPlainText()), 'AQ' + str(row), 10, ws)
            row += 1
            ws.merge_cells('N' + str(row) + ':AB' + str(row))
            self.__cellCenter('(должность)', 'N' + str(row), 8, ws)
            ws.merge_cells('AD' + str(row) + ':AO' + str(row))
            self.__cellCenter('(подпись)', 'AD' + str(row), 8, ws)
            ws.merge_cells('AQ' + str(row) + ':BO' + str(row))
            self.__cellCenter('(расшифровка подписи)', 'AQ' + str(row), 8, ws)
            row += 1

        row += 1
        self.__cellSimple("Решение администрации:", 'A' + str(row), 10, ws)
        ws.merge_cells('A' + str(row + 1) + ':BX' + str(row + 5))
        self.__cellLeft(str(self.adminDecis.toPlainText()), 'A' + str(row  +1), 10, ws)

        wb.save("reportOP8.xlsx")

    def changeOrg(self):
        self.numOrg.setValue(self.arrElemOrg.index(self.cmbOrg.currentText()))

    def changeStruct(self):
        self.numStruct.setValue(self.arrElemStruct.index(self.cmbStruct.currentText()))

    def keyPressEvent(self, event):
        self.chgTabByKey(event)
        if event.key() == QtCore.Qt.Key_N or event.key() == QtCore.Qt.Key_9:
            self.__addRowTblMain()
            self.__addRowTblFight()
            self.__addRowTblLost()

        if event.key() == QtCore.Qt.Key_D or event.key() == QtCore.Qt.Key_0:
            self.__addRowTblCommis()

        if event.key() == QtCore.Qt.Key_Enter or event.key() == QtCore.Qt.Key_Enter - 1:
            self.__updateMainTblRow()
            self.__updateTblLost()
            self.__updateTblFight()
            self.__updateOKPO()

    def __addRowTblCommis(self):
        if self.tabWidget.currentIndex() != self.tabWidget.count() - 1:
            return

        rowPosition = self.tblCommis.rowCount()
        self.tblCommis.insertRow(rowPosition)

        # поля
        txtEdit = QtWidgets.QTextEdit(self.tblCommis)
        self.tblCommis.setCellWidget(rowPosition, 0, txtEdit)

        txtEdit = QtWidgets.QTextEdit(self.tblCommis)
        self.tblCommis.setCellWidget(rowPosition, 1, txtEdit)

        btn = QtWidgets.QPushButton(self.tblCommis)
        btn.setText('Убрать')
        self.tblCommis.setCellWidget(rowPosition, self.tblCommis.columnCount() - 1, btn)
        btn.clicked.connect(
            lambda *args, rowPosition=rowPosition: self.__delRowTblCommis(self.tblCommis.currentRow())
        )

        self.tblCommis.resizeColumnsToContents()

    def __delRowTblCommis(self, row):
        self.tblCommis.removeRow(row)

    def __updateOKPO(self):
        # для организации
        if self.numOrg.value() >= len(self.arrElemOrg):
            self.cmbOrg.setCurrentIndex(0)
            self.numOrg.setValue(0)
        else:
            self.cmbOrg.setCurrentIndex(self.numOrg.value())
        # для структурного подразделения
        if self.numStruct.value() >= len(self.arrElemStruct):
            self.cmbStruct.setCurrentIndex(0)
            self.numStruct.setValue(0)
        else:
            self.cmbStruct.setCurrentIndex(self.numStruct.value())

    def chgTabByKey(self, event):
        if event.key() == QtCore.Qt.Key_1:
            self.tabWidget.setCurrentWidget(self.tabHead)
        elif event.key() == QtCore.Qt.Key_2:
            self.tabWidget.setCurrentWidget(self.tabMain)
        elif event.key() == QtCore.Qt.Key_3:
            self.__updateTblFight()
            self.tabWidget.setCurrentWidget(self.tabFight)
        elif event.key() == QtCore.Qt.Key_4:
            self.__updateTblLost()
            self.tabWidget.setCurrentWidget(self.tabLost)
        elif event.key() == QtCore.Qt.Key_5:
            self.tabWidget.setCurrentWidget(self.tabFut)

    def getCountFightLost(self, row):
        return self.tblFight.cellWidget(row, 1).value() + self.tblLost.cellWidget(row, 1).value()

    def getPriceCountFightLost(self, row, countFightLost = 0):
        if countFightLost == 0:
            countFightLost = self.getCountFightLost(row)
        return self.tblMain.cellWidget(row, 2).value() * countFightLost

    def __updateTblFight(self):
        for i in range(self.tblMain.rowCount()):
            self.tblFight.setItem(i, 0,
                                         QtWidgets.QTableWidgetItem(self.tblMain.cellWidget(i, 0).currentText()))

            self.tblFight.setItem(i, 2,
                                  QtWidgets.QTableWidgetItem(
                                      str(self.tblMain.cellWidget(i, 2).value() * self.tblFight.cellWidget(i, 1).value())))

            countFightLost = self.getCountFightLost(i)
            priceCountFightLost = self.getPriceCountFightLost(i, countFightLost)

            self.tblFight.setItem(i, 4, QtWidgets.QTableWidgetItem(str(countFightLost)))
            self.tblFight.setItem(i, 5, QtWidgets.QTableWidgetItem(str(priceCountFightLost)))

    def __updateTblLost(self):
        for i in range(self.tblMain.rowCount()):
            self.tblLost.setItem(i, 0,
                                  QtWidgets.QTableWidgetItem(self.tblMain.cellWidget(i, 0).currentText()))

            self.tblLost.setItem(i, 2,
                                  QtWidgets.QTableWidgetItem(
                                      str(self.tblMain.cellWidget(i, 2).value() * self.tblLost.cellWidget(i, 1).value())))

            countFightLost = self.getCountFightLost(i)
            priceCountFightLost = self.getPriceCountFightLost(i, countFightLost)

            self.tblLost.setItem(i, 4, QtWidgets.QTableWidgetItem(str(countFightLost)))
            self.tblLost.setItem(i, 5, QtWidgets.QTableWidgetItem(str(priceCountFightLost)))

    def __updateMainTblRow(self):
        self.__updateTblLost()
        for i in range(self.tblMain.rowCount()):
            if self.tblMain.cellWidget(i, 1).value() >= len(self.arrElemName):
                self.tblMain.cellWidget(i, 0).setCurrentIndex(0)
                self.tblMain.cellWidget(i, 1).setValue(0)
            else:
                self.tblMain.cellWidget(i, 0).setCurrentIndex(self.tblMain.cellWidget(i, 1).value())


            countFightLost = self.getCountFightLost(i)
            priceCountFightLost = self.getPriceCountFightLost(i, countFightLost)

            self.tblMain.setItem(i, 4, QtWidgets.QTableWidgetItem(str(countFightLost)))
            self.tblMain.setItem(i, 5, QtWidgets.QTableWidgetItem(str(priceCountFightLost)))

    def __addRowTblMain(self):
        if self.tabWidget.currentIndex() == 0 or self.tabWidget.currentIndex() == self.tabWidget.count() - 1:
            return

        rowPosition = self.tblMain.rowCount()
        self.tblMain.insertRow(rowPosition)

        # поля
        cmbBox = QtWidgets.QComboBox(self.tblMain)
        cmbBox.addItems(self.arrElemName)
        cmbBox.activated[str].connect(
            lambda : self.setCodeByName(rowPosition)
        )
        self.tblMain.setCellWidget(rowPosition, 0, cmbBox)

        spnBox = QtWidgets.QSpinBox(self.tblMain)
        spnBox.setMaximum(self.maxSpnTbl)

        self.tblMain.setCellWidget(rowPosition, 1, spnBox)
        spnBox.setMaximum(self.maxSpnTbl)

        spnBox = QtWidgets.QSpinBox(self.tblMain)
        spnBox.setMaximum(self.maxSpnTbl)

        self.tblMain.setCellWidget(rowPosition, 2, spnBox)

        txtEdit = QtWidgets.QTextEdit(self.tblMain)
        self.tblMain.setCellWidget(rowPosition, 3, txtEdit)

        self.tblMain.setItem(rowPosition, 4, QtWidgets.QTableWidgetItem("0"))
        self.tblMain.setItem(rowPosition, 5, QtWidgets.QTableWidgetItem("0"))

        btn = QtWidgets.QPushButton(self.tblMain)
        btn.setText('Убрать')
        self.tblMain.setCellWidget(rowPosition, self.tblMain.columnCount() - 1, btn)
        btn.clicked.connect(
            lambda *args, rowPosition=rowPosition: self.__delRowTbl(self.tblMain.currentRow())
        )

        self.tblMain.resizeColumnsToContents()

    def __addRowTblLost(self):
        if self.tabWidget.currentIndex() == 0 or self.tabWidget.currentIndex() == self.tabWidget.count() - 1:
            return

        rowPosition = self.tblLost.rowCount()
        self.tblLost.insertRow(rowPosition)

        # поля
        self.tblLost.setItem(rowPosition, 0, QtWidgets.QTableWidgetItem(self.tblMain.cellWidget(rowPosition, 0).currentText()))
        #self.tblMain.setCellWidget(rowPosition, 0, cmbBox)

        spnBox = QtWidgets.QSpinBox(self.tblLost)
        spnBox.setMaximum(self.maxSpnTbl)
        self.tblLost.setCellWidget(rowPosition, 1, spnBox)

        txtEdit = QtWidgets.QTextEdit(self.tblLost)
        self.tblLost.setCellWidget(rowPosition, 3, txtEdit)

        btn = QtWidgets.QPushButton(self.tblLost)
        btn.setText('Убрать')
        self.tblLost.setCellWidget(rowPosition, self.tblLost.columnCount() - 1, btn)
        btn.clicked.connect(
            lambda *args, rowPosition=rowPosition: self.__delRowTbl(self.tblLost.currentRow())
        )

        self.tblLost.resizeColumnsToContents()

    def __addRowTblFight(self):
        if self.tabWidget.currentIndex() == 0 or self.tabWidget.currentIndex() == self.tabWidget.count() - 1:
            return

        rowPosition = self.tblFight.rowCount()
        self.tblFight.insertRow(rowPosition)

        # поля
        self.tblFight.setItem(rowPosition, 0, QtWidgets.QTableWidgetItem(self.tblMain.cellWidget(rowPosition, 0).currentText()))
        # txtEdit = QtWidgets.QTextEdit(self.tblFight)
        # #txtEdit.setEnabled(False)
        # txtEdit.setText(self.tblMain.cellWidget(rowPosition, 0).currentText())
        # self.tblFight.setCellWidget(rowPosition, 0, txtEdit)
        # #self.tblMain.setCellWidget(rowPosition, 0, cmbBox)

        spnBox = QtWidgets.QSpinBox(self.tblFight)
        spnBox.setMaximum(self.maxSpnTbl)
        self.tblFight.setCellWidget(rowPosition, 1, spnBox)

        txtEdit = QtWidgets.QTextEdit(self.tblFight)
        self.tblFight.setCellWidget(rowPosition, 3, txtEdit)


        btn = QtWidgets.QPushButton(self.tblMain)
        btn.setText('Убрать')
        self.tblFight.setCellWidget(rowPosition, self.tblFight.columnCount() - 1, btn)
        btn.clicked.connect(
            lambda *args, rowPosition=rowPosition: self.__delRowTbl(self.tblFight.currentRow())
        )

        self.tblFight.resizeColumnsToContents()

    def __delRowTbl(self, row):
        self.tblMain.removeRow(row)
        self.tblFight.removeRow(row)
        self.tblLost.removeRow(row)

    def setCodeByName(self, row):
        try:
            strVal = self.tblMain.cellWidget(row, 0).currentText()
            self.tblMain.cellWidget(row, 1).setValue(self.arrElemName.index(strVal))
        except Exception:
            print(Exception)

def main():
    app = QtWidgets.QApplication(sys.argv)
    window = MainWnd()
    window.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()